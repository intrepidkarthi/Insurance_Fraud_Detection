import sqlite3
from mimesis import Person
from random import randint, uniform
from faker import Faker
from datetime import date, datetime
import pickle
import calendar
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

'''
    Globals
'''
m_status = ("Married", "Single", "Divorced", "Single", "Single", "Married", "Married")
c_loss = ("Four wheeler accident", "Travel insurance related", "Drive U", "Mobile Repair", "Abhi bus related", "TV repair", "Two wheeler accident", "Fire", "Water", "Theft", "Natural Disaster")
n_loss = ("Borrowed", "Misplaced", "Donated")
i_insurer = ("ola", "bajaj alliance", "icici lombard", "Bharati axa", "coverfox", "future generali","Tata AIG","HDFC ERGO", "United India Insurance", "Reliance", "Universal Sompo")
fraud_reasons = ("No Date of birth", "Date of birth calculated Age and Age do not match",
                 "Claim amount is more than Sum Insured",
                 "No Policy start date", "No Policy end date", "Policy end date before start date",
                 "Claim Date before loss", "No kind of loss", "Invalid kind of loss", "No premium but has claim",
                 "Claim after Policy end date", "Claim before Policy start", "Age is not in requirements")
person = Person('en')
first_name = ('Aadrika',' Aanandinii',' Aaratrika',' Aarya',' Arya',' Aashritha',' Aatmaja',' Atmaja',' Abhaya',' Adwitiya',' Agrata',' Ahilya',' Ahalya',' Aishani',' Akshainie',' Akshata',' Akshita',' Akula',' Ambar',' Amodini',' Amrita',' Amritambu',' Anala',' Anamika',' Ananda',' Anandamayi',' Ananta',' Anila',' Anjali',' Anjushri',' Anjushree',' Annapurna',' Anshula',' Anuja',' Anusuya',' Anasuya',' Anasooya',' Anwesha',' Apsara',' Aruna',' Asha',' Aasa',' Aasha',' Aslesha',' Atreyi',' Atreyee',' Avani',' Abani',' Avantika',' Ayushmati',' Baidehi',' Vaidehi',' Bala',' Baala',' Balamani',' Basanti',' Vasanti',' Bela',' Bhadra',' Bhagirathi',' Bhagwanti',' Bhagwati',' Bhamini',' Bhanumati',' Bhaanumati',' Bhargavi',' Bhavani',' Bhilangana',' Bilwa',' Bilva',' Buddhana',' Chakrika',' Chanda',' Chandi',' Chandni',' Chandini',' Chandani',' Chandra',' Chandira',' Chandrabhaga',' Chandrakala',' Chandrakin',' Chandramani',' Chandrani',' Chandraprabha',' Chandraswaroopa',' Chandravati',' Chapala',' Charumati',' Charvi',' Chatura',' Chitrali',' Chitramala',' Chitrangada',' Daksha',' Dakshayani',' Damayanti',' Darshwana',' Deepali',' Dipali',' Deeptimoyee',' Deeptimayee',' Devangana',' Devani',' Devasree',' Devi',' Daevi',' Devika',' Daevika',' Dhaanyalakshmi',' Dhanalakshmi',' Dhana',' Dhanadeepa',' Dhara',' Dharani',' Dharitri',' Dhatri',' Diksha',' Deeksha',' Divya',' Draupadi',' Dulari',' Durga',' Durgeshwari',' Ekaparnika',' Elakshi',' Enakshi',' Esha',' Eshana',' Eshita',' Gautami',' Gayatri',' Geeta',' Geetanjali',' Gitanjali',' Gemine',' Gemini',' Girja',' Girija',' Gita',' Hamsini',' Harinakshi',' Harita',' Heema',' Himadri',' Himani',' Hiranya',' Indira',' Jaimini',' Jaya',' Jyoti',' Jyotsana',' Kali',' Kalinda',' Kalpana',' Kalyani',' Kama',' Kamala',' Kamla',' Kanchan',' Kanishka',' Kanti',' Kashyapi',' Kumari',' Kumuda',' Lakshmi',' Laxmi',' Lalita',' Lavanya',' Leela',' Lila',' Leela',' Madhuri',' Malti',' Malati',' Mandakini',' Mandaakin',' Mangala',' Mangalya',' Mani',' Manisha',' Manjusha',' Meena',' Mina',' Meenakshi',' Minakshi',' Menka',' Menaka',' Mohana',' Mohini',' Nalini',' Nikita',' Ojaswini',' Omana',' Oormila',' Urmila',' Opalina',' Opaline',' Padma',' Parvati',' Poornima',' Purnima',' Pramila',' Prasanna',' Preity',' Prema',' Priya',' Priyala',' Pushti',' Radha',' Rageswari',' Rageshwari',' Rajinder',' Ramaa',' Rati',' Rita',' Rohana',' Rukhmani',' Rukmin',' Rupinder',' Sanya',' Sarada',' Sharda',' Sarala',' Sarla',' Saraswati',' Sarisha',' Saroja',' Shakti',' Shakuntala',' Shanti',' Sharmila',' Shashi',' Shashikala',' Sheela',' Shivakari',' Shobhana',' Shresth',' Shresthi',' Shreya',' Shreyashi',' Shridevi',' Shrishti',' Shubha',' Shubhaprada',' Siddhi',' Sitara',' Sloka',' Smita',' Smriti',' Soma',' Subhashini',' Subhasini',' Sucheta',' Sudeva',' Sujata',' Sukanya',' Suma',' Suma',' Sumitra',' Sunita',' Suryakantam',' Sushma',' Swara',' Swarnalata',' Sweta',' Shwet',' Tanirika',' Tanushree',' Tanushri',' Tanushri',' Tanya',' Tara',' Trisha',' Uma',' Usha',' Vaijayanti',' Vaijayanthi',' Baijayanti',' Vaishvi',' Vaishnavi',' Vaishno',' Varalakshmi',' Vasudha',' Vasundhara',' Veda',' Vedanshi',' Vidya',' Vimala',' Vrinda',' Vrund',' Aadi',' Aadidev',' Aadinath',' Aaditya',' Aagam',' Aagney',' Aamod',' Aanandaswarup',' Anand Swarup',' Aanjaneya',' Anjaneya',' Aaryan',' Aryan',' Aatmaj',' Aatreya',' Aayushmaan',' Aayushman',' Abhaidev',' Abhaya',' Abhirath',' Abhisyanta',' Acaryatanaya',' Achalesvara',' Acharyanandana',' Acharyasuta',' Achintya',' Achyut',' Adheesh',' Adhiraj',' Adhrit',' Adikavi',' Adinath',' Aditeya',' Aditya',' Adityanandan',' Adityanandana',' Adripathi',' Advaya',' Agasti',' Agastya',' Agneya',' Aagneya',' Agnimitra',' Agniprava',' Agnivesh',' Agrata',' Ajit',' Ajeet',' Akroor',' Akshaj',' Akshat',' Akshayakeerti',' Alok',' Aalok',' Amaranaath',' Amarnath',' Amaresh',' Ambar',' Ameyatma',' Amish',' Amogh',' Amrit',' Anaadi',' Anagh',' Anal',' Anand',' Aanand',' Anang',' Anil',' Anilaabh',' Anilabh',' Anish',' Ankal',' Anunay',' Anurag',' Anuraag',' Archan',' Arindam',' Arjun',' Arnesh',' Arun',' Ashlesh',' Ashok',' Atmanand',' Atmananda',' Avadhesh',' Baalaaditya',' Baladitya',' Baalagopaal',' Balgopal',' Balagopal',' Bahula',' Bakula',' Bala',' Balaaditya',' Balachandra',' Balagovind',' Bandhu',' Bandhul',' Bankim',' Bankimchandra',' Bhadrak',' Bhadraksh',' Bhadran',' Bhagavaan',' Bhagvan',' Bharadwaj',' Bhardwaj',' Bharat',' Bhargava',' Bhasvan',' Bhaasvan',' Bhaswar',' Bhaaswar',' Bhaumik',' Bhaves',' Bheeshma',' Bhisham',' Bhishma',' Bhima',' Bhoj',' Bhramar',' Bhudev',' Bhudeva',' Bhupati',' Bhoopati',' Bhoopat',' Bhupen',' Bhushan',' Bhooshan',' Bhushit',' Bhooshit',' Bhuvanesh',' Bhuvaneshwar',' Bilva',' Bodhan',' Brahma',' Brahmabrata',' Brahmanandam',' Brahmaanand',' Brahmdev',' Brajendra',' Brajesh',' Brijesh',' Birjesh',' Budhil',' Chakor',' Chakradhar',' Chakravartee',' Chakravarti',' Chanakya',' Chaanakya',' Chandak',' Chandan',' Chandra',' Chandraayan',' Chandrabhan',' Chandradev',' Chandraketu',' Chandramauli',' Chandramohan',' Chandran',' Chandranath',' Chapal',' Charak',' Charuchandra',' Chaaruchandra',' Charuvrat',' Chatur',' Chaturaanan',' Chaturbhuj',' Chetan',' Chaten',' Chaitan',' Chetanaanand',' Chidaakaash',' Chidaatma',' Chidambar',' Chidambaram',' Chidananda',' Chinmayanand',' Chinmayananda',' Chiranjeev',' Chiranjeeve',' Chitraksh',' Daiwik',' Daksha',' Damodara',' Dandak',' Dandapaani',' Darshan',' Datta',' Dayaamay',' Dayamayee',' Dayaananda',' Dayaanidhi',' Kin',' Deenabandhu',' Deepan',' Deepankar',' Dipankar',' Deependra',' Dipendra',' Deepesh',' Dipesh',' Deeptanshu',' Deeptendu',' Diptendu',' Deeptiman',' Deeptimoy',' Deeptimay',' Dev',' Deb',' Devadatt',' Devagya',' Devajyoti',' Devak',' Devdan',' Deven',' Devesh',' Deveshwar',' Devi',' Devvrat',' Dhananjay',' Dhanapati',' Dhanpati',' Dhanesh',' Dhanu',' Dhanvin',' Dharmaketu',' Dhruv',' Dhyanesh',' Dhyaneshwar',' Digambar',' Digambara',' Dinakar',' Dinkar',' Dinesh',' Divaakar',' Divakar',' Deevakar',' Divjot',' Dron',' Drona',' Dwaipayan',' Dwaipayana',' Eekalabya',' Ekalavya',' Ekaksh',' Ekaaksh',' Ekaling',' Ekdant',' Ekadant',' Gajaadhar',' Gajadhar',' Gajbaahu',' Gajabahu',' Ganak',' Ganaka',' Ganapati',' Gandharv',' Gandharva',' Ganesh',' Gangesh',' Garud',' Garuda',' Gati',' Gatik',' Gaurang',' Gauraang',' Gauranga',' Gouranga',' Gautam',' Gautama',' Goutam',' Ghanaanand',' Ghanshyam',' Ghanashyam',' Giri',' Girik',' Girika',' Girindra',' Giriraaj',' Giriraj',' Girish',' Gopal',' Gopaal',' Gopi',' Gopee',' Gorakhnath',' Gorakhanatha',' Goswamee',' Goswami',' Gotum',' Gautam',' Govinda',' Gobinda',' Gudakesha',' Gudakesa',' Gurdev',' Guru',' Hari',' Harinarayan',' Harit',' Himadri',' Hiranmay',' Hiranmaya',' Hiranya',' Inder',' Indra',' Indra',' Jagadish',' Jagadisha',' Jagathi',' Jagdeep',' Jagdish',' Jagmeet',' Jahnu',' Jai',' Javas',' Jay',' Jitendra',' Jitender',' Jyotis',' Kailash',' Kama',' Kamalesh',' Kamlesh',' Kanak',' Kanaka',' Kannan',' Kannen',' Karan',' Karthik',' Kartik',' Karunanidhi',' Kashyap',' Kiran',' Kirti',' Keerti',' Krishna',' Krishnadas',' Krishnadasa',' Kumar',' Lai',' Lakshman',' Laxman',' Lakshmidhar',' Lakshminath',' Lal',' Laal',' Mahendra',' Mohinder',' Mahesh',' Maheswar',' Mani',' Manik',' Manikya',' Manoj',' Marut',' Mayoor',' Meghnad',' Meghnath',' Mohan',' Mukesh',' Mukul',' Nagabhushanam',' Nanda',' Narayan',' Narendra',' Narinder',' Naveen',' Navin',' Nawal',' Naval',' Nimit',' Niranjan',' Nirbhay',' Niro',' Param',' Paramartha',' Pran',' Pranay',' Prasad',' Prathamesh',' Prayag',' Prem',' Puneet',' Purushottam',' Rahul',' Raj',' Rajan',' Rajendra',' Rajinder',' Rajiv',' Rakesh',' Ramesh',' Rameshwar',' Ranjit',' Ranjeet',' Ravi',' Ritesh',' Rohan',' Rohit',' Rudra',' Sachin',' Sameer',' Samir',' Sanjay',' Sanka',' Sarvin',' Satish',' Satyen',' Shankar',' Shantanu',' Shashi',' Sher',' Shiv',' Siddarth',' Siddhran',' Som',' Somu',' Somnath',' Subhash',' Subodh',' Suman',' Suresh',' Surya',' Suryakant',' Suryakanta',' Sushil',' Susheel',' Swami',' Swapnil',' Tapan',' Tara',' Tarun',' Tej',' Tejas',' Trilochan',' Trilochana',' Trilok',' Trilokesh',' Triloki',' Triloki Nath',' Trilokanath',' Tushar',' Udai',' Udit',' Ujjawal',' Ujjwal',' Umang',' Upendra',' Uttam',' Vasudev',' Vasudeva',' Vedang',' Vedanga',' Vidhya',' Vidur',' Vidhur',' Vijay',' Vimal',' Vinay',' Vishnu',' Bishnu',' Vishwamitra',' Vyas',' Yogendra',' Yoginder',' Yogesh')
last_name = ('Abbott',' Achari',' Acharya',' Adiga',' Agarwal',' Ahluwalia',' Ahuja',' Arora',' Asan',' Bandopadhyay',' Banerjee',' Bharadwaj',' Bhat',' Butt',' Bhattacharya',' Bhattathiri',' Chaturvedi',' Chattopadhyay',' Chopra',' Desai',' Deshpande',' Devar',' Dhawan',' Dubashi',' Dutta',' Dwivedi',' Embranthiri',' Ganaka',' Gandhi',' Gill',' Gowda',' Guha',' Guneta',' Gupta',' Iyer',' Iyengar',' Jain',' Jha',' Johar',' Joshi',' Kakkar',' Kaniyar',' Kapoor',' Kaul',' Kaur',' Khan',' Khanna',' Khatri',' Kocchar',' Mahajan',' Malik',' Marar',' Menon',' Mehra','Mehrotra',' Mishra',' Mukhopadhyay',' Nadar',' Nayar',' Naik',' Nair',' Nambeesan',' Namboothiri',' Nehru',' Pandey',' Panicker',' Patel',' Patil',' Pilla',' Pillai',' Pothuvaal',' Prajapat',' Rana',' Reddy',' Saini',' Sethi',' Shah',' Sharma',' Shukla',' Singh',' Sinha',' Somayaji',' Tagore',' Talwar',' Tandon',' Trivedi',' Varrier',' Varma',' Varman',' Verma')

city = ("Bengaluru", "Chennai", "Hyderabad", "Kolkata", "Mumbai", "New Delhi", "Thiruvananthapuram", "Visakhapatnam", "Madurai", "Mysuru","Jaipur","Chandigarh","Bhubaneshwar","Amaravati","Gurgaon","Noida","Prayagraj","Patna", "Coimbatore","Kannur","Cochin","Hampi","Mandya","Nellur","Tirupati","Nalgonda","Ahemedabad","Bhopal","Nashik","Ranchi")
state = ("Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh", "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jammu and Kashmir", "Jharkhand", "Karnataka", "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram", "Nagaland", "Orissa", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu", "Tripura", "Uttaranchal", "Uttar Pradesh", "West Bengal", "Andaman and Nicobar Islands", "Dadar and Nagar Haveli", "Daman and Diu", "Delhi", "Lakshadweep", "Pondicherry")

fake = Faker()


mindate = datetime.strptime('Jun 1 1900  1:33PM', '%b %d %Y %I:%M%p')
maxdate = datetime.today()


'''
    Functions
'''

'''
    :param n - number of claims to insert
    :param f - number of fraud claims
'''


def create_database_excel(n, f):
    fraud = set([int(randint(0, n)) for i in range(f)])
    pickle.dump(fraud, open("fraud-pickle.txt", "wb"))
    text_file = open("fraud-index.txt", "w")
    text_file.write("%s" % ', '.join(str(e) for e in fraud))
    text_file.close()

    from openpyxl import load_workbook, Workbook
    wb = Workbook()
    ws = wb.active

    ws.append(["Claim_ID","Name","Surname","Age","Gender","Marital_Status","Date_Of_Birth","Sum_Insured","Policies_Revenue","Policy_Start","Policy_End",
        "Fraudulent_Claim","Fraudulent_Claim_Reason","Date_Of_Loss","Date_Of_Claim","Broker_ID","Insured_ID","Kind_Of_Loss","Claim_Amount",
        "Party_Name","Party_Surname","Service_Provider","Policy_Holder_Street","Policy_Holder_Province","Policy_Holder_City",
        "Policy_Holder_Area","Policy_Holder_Postal","Province","City","Area","Postal_Code"])

    for i in range(0, n):
        if i not in fraud:          
            ws.append(get_data(True))
            print("\rInserted: " + str(i), end="")
        else:
            ws.append(get_data(False))
            print("\rInserted: " + str(i), end="")

    print("\nAll data inserted successfully")		
    wb.save("insurance.xlsx")
    print("Created Database table successfully!")


'''
    :param n - number of claims to insert
    :param f - number of fraud claims
'''


def create_database(n, f):
    fraud = set([int(randint(0, n)) for i in range(f)])
    print(fraud)
    pickle.dump(fraud, open("fraud-pickle.txt", "wb"))
    text_file = open("fraud-index.txt", "w")
    text_file.write("%s" % ', '.join(str(e) for e in fraud))
    text_file.close()

    from openpyxl import load_workbook, Workbook
    wb = Workbook()
    ws = wb.active

    ws.append(["Claim_ID", "Name", "Surname", "Age", "Gender", "Marital_Status", "Date_Of_Birth", "Sum_Insured",
               "Policies_Revenue", "Policy_Start", "Policy_End",
               "Fraudulent_Claim", "Fraudulent_Claim_Reason", "Date_Of_Loss", "Date_Of_Claim", "Broker_ID",
               "Insured_ID", "Kind_Of_Loss", "Claim_Amount",
               "Party_Name", "Party_Surname", "Service_Provider", "Policy_Holder_Street", "Policy_Holder_Province",
               "Policy_Holder_City",
               "Policy_Holder_Area", "Policy_Holder_Postal", "Province", "City", "Area", "Postal_Code"])

    conn = sqlite3.connect('insurance.db')
    cur = conn.cursor()
    print("Opened database successfully")

    cur.execute('''CREATE TABLE IF NOT EXISTS Claims
           (Claim_ID                INTEGER PRIMARY KEY  AUTOINCREMENT NOT NULL,
           Name                     TEXT,
           Surname                  TEXT,
           Age                      INT,
           Gender                   VARCHAR(8),
           Marital_Status           TEXT,
           Date_Of_Birth            DATE,
           Sum_Insured              REAL,
           Policies_Revenue         REAL,
           Policy_Start             DATE,
           Policy_End               DATE,
           Fraudulent_Claim         VARCHAR(1),
           Fraudulent_Claim_Reason  TEXT,
           Date_Of_Loss             DATE,
           Date_Of_Claim            DATE,
           Broker_ID                TEXT,
           Insured_ID               TEXT,
           Kind_Of_Loss             TEXT,
           Claim_Amount             REAL,
           Party_Name               TEXT,
           Party_Surname            TEXT,
           Service_Provider         TEXT,
           Policy_Holder_Street     TEXT,
           Policy_Holder_Province   TEXT,
           Policy_Holder_City       TEXT,
           Policy_Holder_Area       TEXT,
           Policy_Holder_Postal     TEXT,
           Province                 TEXT,
           City                     TEXT,
           Area                     TEXT,
           Postal_Code              TEXT);''')
    print("Created Database table successfully!")

    for i in range(0, n):
        if i not in fraud:
            ws.append(get_data(True))
            cur.execute("INSERT INTO Claims VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        get_data(True))
            print("\rInserted: " + str(i), end="")
        else:
            ws.append(get_data(False))
            cur.execute("INSERT INTO Claims VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        get_data(False))
            print("\rInserted: " + str(i), end="")


    wb.save("insurance.xlsx")
    print("\nAll data inserted successfully in xls sheet")
    conn.commit()
    conn.close()

    print("\nAll data inserted successfully")

'''
    :param status - sends the function to generate a fraud claim or valid claim
'''


def get_data(status):
    dob = random_date()
    dateloss = rand_date("-40y", "now")
    policystart = rand_date("-40y", "now")
    suminsured = random_real(100000, 5000000)

    dobiso = dob.isoformat()
    policystartiso = policystart.isoformat()
    datelossiso = dateloss.isoformat()
    policyend = policy_end(policystart, True)
    if policyend != None:
        policyendiso = policyend.isoformat()
    else:
        policyendiso = None

    if not status:
        return get_fraud_data()

    return (
        null_val(),
        first_name[randint(0, len(first_name) - 1)],
        last_name[randint(0, len(last_name) - 1)],
        calculate_age(dob, status),
        person.gender(),
        marital_status(),
        dobiso,
        suminsured,
        random_real(100, 5000),
        policystartiso,
        policyendiso,
        "F", "",
        datelossiso,
        date_claim(dateloss, policystart, policyend, True),
        "BKR" + str(randint(1000, 9999)), i_insurer[randint(0, len(i_insurer) - 1)],
        c_loss[randint(0, len(c_loss) - 1)],
        claim_amount(suminsured, True),
        first_name[randint(0, len(first_name) - 1)],
        last_name[randint(0, len(last_name) - 1)],
        fake.company(),
        fake.street_name(),
        "India",
        city[randint(0, len(city) - 1)],
        state[randint(0, len(state) - 1)],
        randint(100001, 700000),

        "India",
        city[randint(0, len(city) - 1)],
        state[randint(0, len(state) - 1)],
        randint(100001, 700000)
    )

'''
    function to generate fake data based on the status - False to the helper functions
'''


def get_fraud_data():
    dob = random_date()
    dateloss = rand_date("-40y", "now")
    policystart = rand_date("-40y", "now")

    suminsured = random_real(100000, 5000000)
    r = randint(1, len(fraud_reasons)-1)
    dobiso = dob.isoformat()
    policystartiso = policystart.isoformat()

    datelossiso = dateloss.isoformat()

    if r == 1:
        dob = ""
        dobiso = ""

    if r == 4:
        policystart = ""
        policystartiso = ""

    policyend = policy_end(policystart, r)

    if policyend != None:
        policyendiso = policyend.isoformat()
    else:
        policyendiso = None

    return (
        null_val(),
        first_name[randint(0, len(first_name) - 1)],
        last_name[randint(0, len(last_name) - 1)],
        calculate_age(dob, r),
        person.gender(),
        marital_status(),
        dobiso,
        suminsured,
        premium(r),
        policystartiso,
        policyendiso,
        "T", fraud_reasons[r - 1],
        datelossiso,
        date_claim(dateloss, policystart, policyend, r),
        "BKR" + str(randint(1000, 9999)), i_insurer[randint(0, len(i_insurer) - 1)],
        c_loss[randint(0, len(c_loss) - 1)],
        claim_amount(suminsured, r),
        first_name[randint(0, len(first_name) - 1)],
        last_name[randint(0, len(last_name) - 1)],
        fake.company(),
        fake.street_name(),
        "India",
        city[randint(0, len(city) - 1)],
        state[randint(0, len(state) - 1)],
        randint(100001, 700000),

        "India",
        city[randint(0, len(city) - 1)],
        state[randint(0, len(state) - 1)],
        randint(100001, 700000)
    )


def premium(s):
    if s == 10:
        return 0
    return random_real(100, 5000)


def kind_loss(s):
    if s == 8:
        return None
    if s == 9:
        return n_loss[randint(0, len(c_loss) - 1)]
    return c_loss[randint(0, len(c_loss) - 1)]


def date_claim(loss, policystart, policyend, s):
    if loss == "":
        return None
    if policyend == "":
        return None
    if policystart == "":
        return None
    if s == 7:
        return date_between(mindate, loss).isoformat()
    if s == 11:
        return date_between(policyend, maxdate)
    if s == 12:
        return date_between(mindate, policystart)
    return date_between(loss, maxdate).isoformat()


def policy_end(start, s):
    if start == "":
        return None
    if s == 6:
        return date_between(mindate, start)
    if s == 5:
        return None
    return date_between(start, maxdate)

"""
    This function will randomly generate a date between start and end dates provided
    :param s - start date 
    :param e - end date
"""


def date_between(s, e):
    y = randint(s.year, e.year)
    m = randint(1, 12)
    d = randint(1, 30)

    if calendar.isleap(y):
        if m == 2:
            d = randint(1, 29)

    if m == 2:
        d = randint(1, 28)

    h = randint(0, 12)
    i = randint(0, 59)
    s = randint(0, 59)

    return datetime(y, m, d, h, i, s)


def claim_amount(val, s):
    if s == 3:
        r = randint(0, 100)
        if r < 60:
            return random_real(val, 9000)
        if r > 60 & r < 80:
            return random_real(val, 50000)
        if r > 80 & r < 90:
            return random_real(val, 90000)
        if r > 90:
            return random_real(val, 900000)
    return random_real(1, val)

"""
    This function will calculate the age depending on the provided date of birth and status, where the status 
    will determine if a actual or fraudulent age must be created.
    
    :param born - date of birth
    :param s - status
"""


def calculate_age(born, s):
    if s == 2:
        return person.age()
    if s == 4:
        return person.age()
    if s == 13:
        if randint(0, 1) == 0:
            return randint(-10, 15)
        else:
            return randint(120, 300)
    if born == "":
        return None
    today = date.today()
    return today.year - born.year - ((today.month, today.day) < (born.month, born.day))


def marital_status():
    return m_status[randint(0, len(m_status) - 1)]


"""
    This function will generate a random date between 1920 and 1999, which is all validated and caters for leap years 
"""


def random_date():
    y = randint(1920, 2001)
    m = randint(1, 12)
    d = randint(1, 30)
    if calendar.isleap(y):
        if m == 2:
            d = randint(1, 29)

    if m == 2:
        d = randint(1, 28)

    h = randint(0, 12)
    i = randint(0, 59)
    s = randint(0, 59)

    return datetime(y, m, d, h, i, s)


def rand_date(start, end):
    return fake.date_time_between(start, end)


def random_real(m, mm):
    return round(uniform(m, mm), 2)


def null_val():
    return None


'''
    SCRIPT
    
    eg. create_database(number of claims, number of fraud claims)
'''

start_time = time.time()
create_database(10000, 1456)
print("--- %s seconds ---" % (time.time() - start_time))

'''
    Data Cleaning
        - Check if DOB and age is correct
        - check claim date
        - policy expire
        - amount claim vs insured
        - check empty cells
'''
